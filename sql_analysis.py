"""
==============================================================================
 DecodeLabs - Data Analytics Project 3: SQL Data Analysis
 Batch: 2026
 Dataset: Dataset for Data Analytics (3).xlsx
 
 Goal: Use SQL queries to extract insights from an e-commerce orders dataset.
 Key Skills: SQL fundamentals, querying data, filtering, grouping, aggregations
==============================================================================
"""

import sqlite3
import openpyxl
import os
import sys
import io
from datetime import datetime

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ─────────────────────────────────────────────────────────────
# STEP 1: Load Excel Data into SQLite Database
# ─────────────────────────────────────────────────────────────

def load_data_to_sqlite(excel_file, db_file="ecommerce_orders.db"):
    """Load Excel dataset into a SQLite database table."""
    
    # Remove existing DB to start fresh
    if os.path.exists(db_file):
        os.remove(db_file)
    
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    
    # Create the Orders table
    cursor.execute("""
        CREATE TABLE Orders (
            OrderID TEXT PRIMARY KEY,
            Date TEXT,
            CustomerID TEXT,
            Product TEXT,
            Quantity INTEGER,
            UnitPrice REAL,
            ShippingAddress TEXT,
            PaymentMethod TEXT,
            OrderStatus TEXT,
            TrackingNumber TEXT,
            ItemsInCart INTEGER,
            CouponCode TEXT,
            ReferralSource TEXT,
            TotalPrice REAL
        );
    """)
    
    # Load data from Excel
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date_val = row[1].strftime('%Y-%m-%d') if isinstance(row[1], datetime) else str(row[1])
        coupon = row[11] if row[11] else None
        
        cursor.execute("""
            INSERT INTO Orders VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            row[0], date_val, row[2], row[3], row[4], row[5],
            row[6], row[7], row[8], row[9], row[10],
            coupon, row[12], row[13]
        ))
        row_count += 1
    
    conn.commit()
    print(f"✅ Successfully loaded {row_count} records into SQLite database '{db_file}'")
    print(f"   Table: Orders | Columns: 14")
    print()
    
    return conn


def run_query(conn, query_number, title, sql, description=""):
    """Execute a SQL query and display formatted results."""
    
    print("=" * 80)
    print(f"  QUERY {query_number}: {title}")
    print("=" * 80)
    
    if description:
        print(f"  Purpose: {description}")
    
    print(f"\n  SQL:\n  {sql.strip()}\n")
    print("-" * 80)
    
    cursor = conn.cursor()
    cursor.execute(sql)
    
    # Get column names
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    
    # Calculate column widths
    col_widths = []
    for i, col in enumerate(columns):
        max_width = len(str(col))
        for row in rows:
            max_width = max(max_width, len(str(row[i])))
        col_widths.append(min(max_width, 30))  # Cap at 30 chars
    
    # Print header
    header = " | ".join(str(col).ljust(col_widths[i]) for i, col in enumerate(columns))
    print(f"  {header}")
    print(f"  {'-' * len(header)}")
    
    # Print rows
    for row in rows:
        row_str = " | ".join(str(val).ljust(col_widths[i])[:30] for i, val in enumerate(row))
        print(f"  {row_str}")
    
    print(f"\n  → {len(rows)} row(s) returned")
    print()
    
    return rows


def main():
    """Main function to run all SQL queries."""
    
    print()
    print("╔" + "═" * 78 + "╗")
    print("║" + " DecodeLabs - Project 3: SQL Data Analysis".center(78) + "║")
    print("║" + " E-Commerce Orders Dataset Analysis".center(78) + "║")
    print("║" + " Batch 2026 | 1,200 Orders | 7 Products".center(78) + "║")
    print("╚" + "═" * 78 + "╝")
    print()
    
    # Load data
    excel_file = "Dataset for Data Analytics (3).xlsx"
    conn = load_data_to_sqlite(excel_file)
    
    # ─────────────────────────────────────────────────────────────
    # SECTION A: SELECT Queries (Basic Data Retrieval)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION A: SELECT QUERIES - Basic Data Retrieval")
    print("█" * 80 + "\n")
    
    # Query 1: Select all columns - first 10 records
    run_query(conn, 1, "View First 10 Orders (SELECT *)",
        """
        SELECT * FROM Orders LIMIT 10;
        """,
        "Retrieve all columns for the first 10 orders to preview the dataset structure."
    )
    
    # Query 2: Select specific columns
    run_query(conn, 2, "Select Specific Columns (OrderID, Product, TotalPrice)",
        """
        SELECT OrderID, Product, Quantity, UnitPrice, TotalPrice
        FROM Orders
        LIMIT 10;
        """,
        "Retrieve only the relevant pricing columns to focus on order values."
    )
    
    # Query 3: Select distinct products
    run_query(conn, 3, "All Unique Products (SELECT DISTINCT)",
        """
        SELECT DISTINCT Product
        FROM Orders
        ORDER BY Product;
        """,
        "Identify all unique products sold in the e-commerce store."
    )
    
    # Query 4: Select distinct payment methods
    run_query(conn, 4, "All Payment Methods Available (SELECT DISTINCT)",
        """
        SELECT DISTINCT PaymentMethod
        FROM Orders
        ORDER BY PaymentMethod;
        """,
        "List all unique payment methods used by customers."
    )
    
    # ─────────────────────────────────────────────────────────────
    # SECTION B: WHERE Clause (Filtering Data)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION B: WHERE CLAUSE - Filtering Data")
    print("█" * 80 + "\n")
    
    # Query 5: Filter by Order Status
    run_query(conn, 5, "All Delivered Orders",
        """
        SELECT OrderID, Product, CustomerID, TotalPrice, OrderStatus
        FROM Orders
        WHERE OrderStatus = 'Delivered'
        LIMIT 15;
        """,
        "Filter orders to show only those that have been successfully delivered."
    )
    
    # Query 6: Filter by Product and Price
    run_query(conn, 6, "Laptop Orders Over $500",
        """
        SELECT OrderID, Product, Quantity, UnitPrice, TotalPrice
        FROM Orders
        WHERE Product = 'Laptop' AND TotalPrice > 500
        ORDER BY TotalPrice DESC
        LIMIT 10;
        """,
        "Find high-value laptop orders exceeding $500 total price."
    )
    
    # Query 7: Filter using IN operator
    run_query(conn, 7, "Orders Paid via Credit Card or Debit Card",
        """
        SELECT OrderID, Product, PaymentMethod, TotalPrice
        FROM Orders
        WHERE PaymentMethod IN ('Credit Card', 'Debit Card')
        LIMIT 15;
        """,
        "Filter orders made using card-based payment methods."
    )
    
    # Query 8: Filter using BETWEEN for dates
    run_query(conn, 8, "Orders Placed in January 2024",
        """
        SELECT OrderID, Date, Product, TotalPrice
        FROM Orders
        WHERE Date BETWEEN '2024-01-01' AND '2024-01-31'
        ORDER BY Date;
        """,
        "Retrieve all orders placed during January 2024."
    )
    
    # Query 9: Filter cancelled or returned orders
    run_query(conn, 9, "Cancelled and Returned Orders (Problem Orders)",
        """
        SELECT OrderID, Product, OrderStatus, TotalPrice, PaymentMethod
        FROM Orders
        WHERE OrderStatus IN ('Cancelled', 'Returned')
        ORDER BY TotalPrice DESC
        LIMIT 15;
        """,
        "Identify problematic orders that were cancelled or returned."
    )
    
    # Query 10: Filter using LIKE for pattern matching
    run_query(conn, 10, "Orders with SAVE10 Coupon",
        """
        SELECT OrderID, Product, CouponCode, TotalPrice
        FROM Orders
        WHERE CouponCode LIKE 'SAVE%'
        LIMIT 10;
        """,
        "Find orders where customers used a SAVE-type coupon code."
    )
    
    # ─────────────────────────────────────────────────────────────
    # SECTION C: ORDER BY (Sorting Data)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION C: ORDER BY - Sorting Data")
    print("█" * 80 + "\n")
    
    # Query 11: Order by TotalPrice Descending (Top orders)
    run_query(conn, 11, "Top 10 Highest Value Orders",
        """
        SELECT OrderID, Product, Quantity, UnitPrice, TotalPrice
        FROM Orders
        ORDER BY TotalPrice DESC
        LIMIT 10;
        """,
        "Find the 10 most expensive orders by total price."
    )
    
    # Query 12: Order by Date (Most Recent)
    run_query(conn, 12, "10 Most Recent Orders",
        """
        SELECT OrderID, Date, Product, CustomerID, TotalPrice
        FROM Orders
        ORDER BY Date DESC
        LIMIT 10;
        """,
        "Retrieve the 10 most recently placed orders."
    )
    
    # Query 13: Multi-column sort
    run_query(conn, 13, "Orders Sorted by Product (A-Z) then Price (High to Low)",
        """
        SELECT OrderID, Product, UnitPrice, Quantity, TotalPrice
        FROM Orders
        ORDER BY Product ASC, TotalPrice DESC
        LIMIT 15;
        """,
        "Sort orders alphabetically by product, then by price within each product."
    )
    
    # ─────────────────────────────────────────────────────────────
    # SECTION D: GROUP BY with Aggregations (COUNT, SUM, AVG)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION D: GROUP BY with AGGREGATIONS (COUNT, SUM, AVG)")
    print("█" * 80 + "\n")
    
    # Query 14: COUNT - Orders per product
    run_query(conn, 14, "Total Number of Orders per Product (COUNT)",
        """
        SELECT Product,
               COUNT(*) AS TotalOrders
        FROM Orders
        GROUP BY Product
        ORDER BY TotalOrders DESC;
        """,
        "Count how many orders were placed for each product category."
    )
    
    # Query 15: SUM - Total revenue per product
    run_query(conn, 15, "Total Revenue per Product (SUM)",
        """
        SELECT Product,
               SUM(TotalPrice) AS TotalRevenue,
               SUM(Quantity) AS TotalUnitsSold
        FROM Orders
        GROUP BY Product
        ORDER BY TotalRevenue DESC;
        """,
        "Calculate total revenue and units sold for each product."
    )
    
    # Query 16: AVG - Average order value per product
    run_query(conn, 16, "Average Order Value per Product (AVG)",
        """
        SELECT Product,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue,
               ROUND(AVG(UnitPrice), 2) AS AvgUnitPrice,
               ROUND(AVG(Quantity), 2) AS AvgQuantity
        FROM Orders
        GROUP BY Product
        ORDER BY AvgOrderValue DESC;
        """,
        "Calculate the average order value, unit price, and quantity for each product."
    )
    
    # Query 17: Orders per Payment Method
    run_query(conn, 17, "Order Count and Revenue by Payment Method",
        """
        SELECT PaymentMethod,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalRevenue,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue
        FROM Orders
        GROUP BY PaymentMethod
        ORDER BY TotalRevenue DESC;
        """,
        "Analyze how each payment method contributes to overall orders and revenue."
    )
    
    # Query 18: Orders by Status
    run_query(conn, 18, "Order Distribution by Status",
        """
        SELECT OrderStatus,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalValue,
               ROUND(100.0 * COUNT(*) / (SELECT COUNT(*) FROM Orders), 2) AS PercentageOfTotal
        FROM Orders
        GROUP BY OrderStatus
        ORDER BY OrderCount DESC;
        """,
        "Show how orders are distributed across different statuses with percentage contribution."
    )
    
    # Query 19: Monthly Revenue Analysis
    run_query(conn, 19, "Monthly Revenue Analysis (2024)",
        """
        SELECT SUBSTR(Date, 1, 7) AS Month,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS MonthlyRevenue,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue
        FROM Orders
        WHERE Date BETWEEN '2024-01-01' AND '2024-12-31'
        GROUP BY SUBSTR(Date, 1, 7)
        ORDER BY Month;
        """,
        "Track monthly order volume and revenue trends throughout 2024."
    )
    
    # Query 20: Revenue by Referral Source
    run_query(conn, 20, "Revenue by Referral Source (Marketing Channel Analysis)",
        """
        SELECT ReferralSource,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalRevenue,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue,
               ROUND(100.0 * SUM(TotalPrice) / (SELECT SUM(TotalPrice) FROM Orders), 2) AS RevenueSharePct
        FROM Orders
        GROUP BY ReferralSource
        ORDER BY TotalRevenue DESC;
        """,
        "Evaluate which marketing channels drive the most orders and revenue."
    )
    
    # ─────────────────────────────────────────────────────────────
    # SECTION E: HAVING Clause (Filtering Grouped Data)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION E: HAVING CLAUSE - Filtering Grouped Data")
    print("█" * 80 + "\n")
    
    # Query 21: Products with average order value > $800
    run_query(conn, 21, "Products with Average Order Value Over $800 (HAVING)",
        """
        SELECT Product,
               COUNT(*) AS OrderCount,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue
        FROM Orders
        GROUP BY Product
        HAVING AVG(TotalPrice) > 800
        ORDER BY AvgOrderValue DESC;
        """,
        "Find high-value product categories where the average order exceeds $800."
    )
    
    # Query 22: Customers with more than 3 orders
    run_query(conn, 22, "Repeat Customers with More Than 3 Orders (HAVING)",
        """
        SELECT CustomerID,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalSpent,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue
        FROM Orders
        GROUP BY CustomerID
        HAVING COUNT(*) > 3
        ORDER BY TotalSpent DESC
        LIMIT 15;
        """,
        "Identify loyal customers who have placed more than 3 orders."
    )
    
    # Query 23: Payment methods with total revenue > $200,000
    run_query(conn, 23, "High-Revenue Payment Methods (HAVING with SUM)",
        """
        SELECT PaymentMethod,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalRevenue
        FROM Orders
        GROUP BY PaymentMethod
        HAVING SUM(TotalPrice) > 200000
        ORDER BY TotalRevenue DESC;
        """,
        "Filter payment methods that have generated more than $200,000 in total revenue."
    )
    
    # ─────────────────────────────────────────────────────────────
    # SECTION F: Advanced Queries (Combined Techniques)
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SECTION F: ADVANCED QUERIES - Combined Techniques")
    print("█" * 80 + "\n")
    
    # Query 24: Coupon Code Effectiveness
    run_query(conn, 24, "Coupon Code Effectiveness Analysis",
        """
        SELECT COALESCE(CouponCode, 'No Coupon') AS Coupon,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalRevenue,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue
        FROM Orders
        GROUP BY CouponCode
        ORDER BY TotalRevenue DESC;
        """,
        "Compare the effectiveness of different coupon codes vs. no coupon."
    )
    
    # Query 25: Year-over-Year Revenue Comparison
    run_query(conn, 25, "Year-over-Year Revenue Comparison",
        """
        SELECT SUBSTR(Date, 1, 4) AS Year,
               COUNT(*) AS TotalOrders,
               ROUND(SUM(TotalPrice), 2) AS TotalRevenue,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue,
               SUM(Quantity) AS TotalUnitsSold
        FROM Orders
        GROUP BY SUBSTR(Date, 1, 4)
        ORDER BY Year;
        """,
        "Compare business performance across different years."
    )
    
    # Query 26: Product Performance by Order Status
    run_query(conn, 26, "Product Performance by Order Status",
        """
        SELECT Product,
               OrderStatus,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalValue
        FROM Orders
        GROUP BY Product, OrderStatus
        ORDER BY Product, OrderCount DESC;
        """,
        "Cross-tabulate products against order statuses to identify problem areas."
    )
    
    # Query 27: Top Spending Customers
    run_query(conn, 27, "Top 10 Highest Spending Customers",
        """
        SELECT CustomerID,
               COUNT(*) AS OrderCount,
               ROUND(SUM(TotalPrice), 2) AS TotalSpent,
               ROUND(AVG(TotalPrice), 2) AS AvgOrderValue,
               MAX(Date) AS LastOrderDate
        FROM Orders
        GROUP BY CustomerID
        ORDER BY TotalSpent DESC
        LIMIT 10;
        """,
        "Identify the top 10 customers by total spending for VIP targeting."
    )
    
    # Query 28: Cancelled Order Rate by Product
    run_query(conn, 28, "Cancellation Rate by Product",
        """
        SELECT Product,
               COUNT(*) AS TotalOrders,
               SUM(CASE WHEN OrderStatus = 'Cancelled' THEN 1 ELSE 0 END) AS CancelledOrders,
               ROUND(100.0 * SUM(CASE WHEN OrderStatus = 'Cancelled' THEN 1 ELSE 0 END) / COUNT(*), 2) AS CancellationRate
        FROM Orders
        GROUP BY Product
        ORDER BY CancellationRate DESC;
        """,
        "Calculate the cancellation rate for each product to identify quality issues."
    )
    
    # ─────────────────────────────────────────────────────────────
    # Summary Statistics
    # ─────────────────────────────────────────────────────────────
    
    print("\n" + "█" * 80)
    print("  SUMMARY: Overall Dataset Statistics")
    print("█" * 80 + "\n")
    
    run_query(conn, "SUMMARY", "Overall Business Metrics",
        """
        SELECT 
               COUNT(*) AS TotalOrders,
               COUNT(DISTINCT CustomerID) AS UniqueCustomers,
               COUNT(DISTINCT Product) AS UniqueProducts,
               ROUND(SUM(TotalPrice), 2) AS GrandTotalRevenue,
               ROUND(AVG(TotalPrice), 2) AS OverallAvgOrderValue,
               MIN(Date) AS FirstOrderDate,
               MAX(Date) AS LastOrderDate
        FROM Orders;
        """,
        "High-level summary of the entire e-commerce orders dataset."
    )
    
    # Close connection
    conn.close()
    
    print()
    print("╔" + "═" * 78 + "╗")
    print("║" + " ✅ SQL Data Analysis Complete!".center(78) + "║")
    print("║" + " All 28 queries executed successfully.".center(78) + "║")
    print("║" + " Database: ecommerce_orders.db".center(78) + "║")
    print("╚" + "═" * 78 + "╝")
    print()


if __name__ == "__main__":
    main()
